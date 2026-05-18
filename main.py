from fastapi import FastAPI, File, UploadFile, Form, HTTPException
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
import xlrd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from datetime import datetime, time, date, timedelta
import calendar
import tempfile
import os
import re
import smtplib
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

try:
    from dotenv import load_dotenv

    load_dotenv(override=True)
except ImportError:
    def _load_env_file(path: str = ".env"):
        if not os.path.isfile(path):
            return
        with open(path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, _, value = line.partition("=")
                key, value = key.strip(), value.strip().strip("'\"")
                if key:
                    os.environ[key] = value

    _load_env_file()

app = FastAPI(title="School Attendance System")
app.mount("/static", StaticFiles(directory="static"), name="static")


@app.get("/", response_class=HTMLResponse)
async def root():
    with open("static/index.html", "r") as f:
        return f.read()


@app.get("/email-status")
async def email_status():
    """Check whether SMTP is configured (for UI hints)."""
    return {"configured": is_email_configured()}


@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    """Parse uploaded XLS file and return teacher list with record counts."""
    content = await file.read()
    try:
        records = parse_xls(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    teachers = {}
    all_dates = []
    for r in records:
        uid = r["user_id"]
        if uid not in teachers:
            teachers[uid] = {"user_id": uid, "name": r["name"], "count": 0}
        teachers[uid]["count"] += 1
        all_dates.append(r["datetime"])

    month_label = ""
    if all_dates:
        d = min(all_dates)
        month_label = d.strftime("%B %Y")

    return {
        "teachers": list(teachers.values()),
        "records": [
            {
                "user_id": r["user_id"],
                "name": r["name"],
                "datetime": r["datetime"].isoformat(),
                "status": r["status"],
            }
            for r in records
        ],
        "total_records": len(records),
        "month": month_label,
    }


@app.post("/generate")
async def generate_report(
    file: UploadFile = File(...),
    school_time: str = Form("08:30"),
    staff_timing: str = Form("08:15"),
    relaxation_minutes: int = Form(5),
    working_days: str = Form("0,1,2,3,4"),
    monthly_leave: int = Form(1),
    off_days: str = Form(""),
):
    """Process attendance data and return generated Excel file."""
    content = await file.read()
    try:
        records = parse_xls(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    settings = parse_settings(
        school_time, staff_timing, relaxation_minutes, working_days, monthly_leave, off_days
    )
    output_path, fname = build_report_xlsx(records, settings)

    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=fname,
        headers={"Access-Control-Expose-Headers": "Content-Disposition"},
    )


@app.post("/send-report")
async def send_teacher_report(
    file: UploadFile = File(...),
    user_id: str = Form(...),
    email: str = Form(...),
    school_time: str = Form("08:30"),
    staff_timing: str = Form("08:15"),
    relaxation_minutes: int = Form(5),
    working_days: str = Form("0,1,2,3,4"),
    monthly_leave: int = Form(1),
    off_days: str = Form(""),
):
    """Generate a single-teacher Excel report and email it."""
    email = email.strip()
    if not _is_valid_email(email):
        raise HTTPException(status_code=400, detail="Please enter a valid email address.")

    content = await file.read()
    try:
        records = parse_xls(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    settings = parse_settings(
        school_time, staff_timing, relaxation_minutes, working_days, monthly_leave, off_days
    )

    try:
        output_path, fname, teacher_name = build_report_xlsx(records, settings, user_id=user_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

    try:
        send_report_email(email, teacher_name, fname, output_path)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to send email: {e}")
    finally:
        if os.path.exists(output_path):
            os.unlink(output_path)

    return {"ok": True, "message": f"Report sent to {email}"}


# ─── Core Logic ───────────────────────────────────────────────────────────────

def parse_settings(
    school_time: str,
    staff_timing: str,
    relaxation_minutes: int,
    working_days: str,
    monthly_leave: int,
    off_days: str,
) -> dict:
    wd_list = [int(x) for x in working_days.split(",") if x.strip()]
    off_days_set = {d.strip() for d in off_days.split(",") if d.strip()}
    return {
        "school_time": school_time,
        "staff_timing": staff_timing,
        "relaxation_minutes": relaxation_minutes,
        "working_days": wd_list,
        "monthly_leave": monthly_leave,
        "off_days": off_days_set,
    }


def _sanitize_filename_part(name: str) -> str:
    cleaned = re.sub(r"[^\w\s-]", "", name, flags=re.UNICODE)
    cleaned = re.sub(r"\s+", "-", cleaned.strip())
    return cleaned or "Teacher"


def _is_valid_email(email: str) -> bool:
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email))


def build_report_xlsx(records: list, settings: dict, user_id: str | None = None) -> tuple:
    """Build Excel report. Returns (output_path, filename[, teacher_name])."""
    teachers_data, national_holidays = calculate_stats(records, settings)

    uid_order = []
    seen = set()
    for r in records:
        if r["user_id"] not in seen:
            uid_order.append(r["user_id"])
            seen.add(r["user_id"])

    if user_id is not None:
        if user_id not in teachers_data:
            raise ValueError(f"Teacher with ID {user_id} not found in uploaded file.")
        ordered_teachers = [teachers_data[user_id]]
        teacher_name = ordered_teachers[0]["name"]
    else:
        ordered_teachers = [teachers_data[uid] for uid in uid_order if uid in teachers_data]
        teacher_name = None

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        output_path = tmp.name

    generate_excel(ordered_teachers, settings, national_holidays, output_path)

    if records:
        d = min(r["datetime"] for r in records)
        month_part = d.strftime("%b-%Y")
    else:
        month_part = "Report"

    if teacher_name:
        safe_name = _sanitize_filename_part(teacher_name)
        fname = f"Attendance-{safe_name}-{month_part}.xlsx"
    else:
        fname = f"Attendance-{month_part}.xlsx"

    if user_id is not None:
        return output_path, fname, teacher_name
    return output_path, fname


def _smtp_password_normalized(host: str, password: str) -> str:
    """Gmail app passwords must be entered without spaces for SMTP."""
    p = password.strip()
    if "gmail" in host.lower():
        return re.sub(r"\s+", "", p)
    return p


def is_email_configured() -> bool:
    host = os.environ.get("SMTP_HOST", "").strip()
    user = os.environ.get("SMTP_USER", "").strip()
    password = os.environ.get("SMTP_PASSWORD", "").strip()
    return bool(host and user and password)


def _email_not_configured_message() -> str:
    return (
        "Email is not configured. "
        "Local: copy .env.example to .env and set SMTP_HOST, SMTP_USER, SMTP_PASSWORD, then restart the server. "
        "Render: add those variables under Environment and redeploy."
    )


def _smtp_auth_error_message(exc: Exception, host: str) -> str:
    err = str(exc).lower()
    if "535" in err or "badcredentials" in err or "authentication failed" in err:
        if "gmail" in host.lower():
            return (
                "Gmail rejected the login (535). Use an App Password (16 characters), not your normal "
                "Gmail password. Paste it without spaces. SMTP_USER must be the same Gmail address. "
                "If App passwords are disabled for your account, use Outlook SMTP or a provider like Resend."
            )
        return (
            "The mail server rejected the username or password (535). "
            "Check SMTP_USER and SMTP_PASSWORD; many providers require an app-specific password, not your web login."
        )
    return f"Mail server error: {exc}"


def send_report_email(to_email: str, teacher_name: str, attachment_name: str, file_path: str):
    host = os.environ.get("SMTP_HOST", "").strip()
    user = os.environ.get("SMTP_USER", "").strip()
    password = _smtp_password_normalized(host, os.environ.get("SMTP_PASSWORD", ""))
    from_addr = os.environ.get("SMTP_FROM", user).strip() or user

    if not is_email_configured():
        raise HTTPException(status_code=503, detail=_email_not_configured_message())

    port = int(os.environ.get("SMTP_PORT", "587"))
    use_tls = os.environ.get("SMTP_USE_TLS", "true").lower() in ("1", "true", "yes")

    subject = f"Attendance Report — {teacher_name}"
    body = (
        f"Dear {teacher_name},\n\n"
        "Please find your attendance report attached.\n\n"
        "Regards,\nSchool Attendance System"
    )

    msg = MIMEMultipart()
    msg["From"] = from_addr
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    with open(file_path, "rb") as f:
        part = MIMEApplication(f.read(), Name=attachment_name)
    part["Content-Disposition"] = f'attachment; filename="{attachment_name}"'
    msg.attach(part)

    try:
        if use_tls:
            with smtplib.SMTP(host, port, timeout=30) as server:
                server.starttls()
                server.login(user, password)
                server.sendmail(from_addr, [to_email], msg.as_string())
        else:
            with smtplib.SMTP_SSL(host, port, timeout=30) as server:
                server.login(user, password)
                server.sendmail(from_addr, [to_email], msg.as_string())
    except smtplib.SMTPAuthenticationError as e:
        raise HTTPException(status_code=401, detail=_smtp_auth_error_message(e, host)) from e
    except smtplib.SMTPException as e:
        raise HTTPException(status_code=502, detail=f"Failed to send email: {e}") from e


def parse_xls(file_content: bytes) -> list:
    wb = xlrd.open_workbook(file_contents=file_content)
    sh = wb.sheet_by_index(0)
    records = []
    datemode = wb.datemode

    for i in range(1, sh.nrows):
        row = [sh.cell_value(i, j) for j in range(sh.ncols)]
        if not row[2]:
            continue
        try:
            user_id = str(int(float(str(row[0])))) if row[0] != "" else ""
            name = str(row[1]).strip()
            dt = xlrd.xldate_as_datetime(row[2], datemode)
            status = str(row[4]).strip() if row[4] else str(row[3]).strip()
            records.append({"user_id": user_id, "name": name, "datetime": dt, "status": status})
        except Exception:
            continue

    return records


def build_cutoff(staff_timing: str, relaxation_minutes: int) -> time:
    h, m = map(int, staff_timing.split(":"))
    total = h * 60 + m + relaxation_minutes
    return time(total // 60, total % 60)


def calculate_stats(records: list, settings: dict):
    working_days_set = set(settings["working_days"])
    cutoff = build_cutoff(settings["staff_timing"], settings["relaxation_minutes"])

    off_days_dates = set()
    for ds in settings.get("off_days", set()):
        try:
            off_days_dates.add(datetime.strptime(ds, "%Y-%m-%d").date())
        except ValueError:
            continue

    # Group by teacher preserving insertion order
    teachers: dict = {}
    for r in records:
        uid = r["user_id"]
        if uid not in teachers:
            teachers[uid] = {"user_id": uid, "name": r["name"], "records": []}
        teachers[uid]["records"].append(r)

    # Determine month range
    all_dates = [r["datetime"].date() for r in records]
    if not all_dates:
        return {}, set()
    year = min(all_dates).year
    month = min(all_dates).month
    num_days = calendar.monthrange(year, month)[1]

    all_working_days = [
        date(year, month, d)
        for d in range(1, num_days + 1)
        if date(year, month, d).weekday() in working_days_set
        and date(year, month, d) not in off_days_dates
    ]

    # Days where at least one teacher has an In record (excluding custom off days)
    days_with_any_in: set = set()
    for teacher in teachers.values():
        for r in teacher["records"]:
            if r["status"].lower() == "in":
                d = r["datetime"].date()
                if d.weekday() in working_days_set and d not in off_days_dates:
                    days_with_any_in.add(d)

    # National holidays: working days with ZERO attendance across all teachers
    national_holidays = set(all_working_days) - days_with_any_in

    for uid, teacher in teachers.items():
        records_by_day: dict = {}
        for r in teacher["records"]:
            d = r["datetime"].date()
            records_by_day.setdefault(d, []).append(r)

        right_time = 0
        late = 0
        late_dates: set = set()

        for d, day_records in records_by_day.items():
            if d.weekday() not in working_days_set or d in off_days_dates:
                continue
            in_records = [r for r in day_records if r["status"].lower() == "in"]
            if not in_records:
                continue
            first_in = min(in_records, key=lambda x: x["datetime"])
            if first_in["datetime"].time() <= cutoff:
                right_time += 1
            else:
                late += 1
                late_dates.add(d)

        days_teacher_attended = {
            d for d in records_by_day
            if d.weekday() in working_days_set
            and d not in off_days_dates
            and any(r["status"].lower() == "in" for r in records_by_day[d])
        }
        leave = len(
            [d for d in all_working_days if d not in national_holidays and d not in days_teacher_attended]
        )

        teacher["right_time"] = right_time
        teacher["late"] = late
        teacher["late_dates"] = late_dates
        teacher["leave"] = leave

    return teachers, national_holidays


# ─── Excel Generation ─────────────────────────────────────────────────────────

def _make_border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
SUBHEADER_FILL = PatternFill("solid", fgColor="2E6DA4")
ALT_ROW_FILL = PatternFill("solid", fgColor="EBF3FB")
LATE_ROW_FILL = PatternFill("solid", fgColor="FFF9C4")
SUMMARY_FILL = PatternFill("solid", fgColor="E8F5E9")
IN_FONT_COLOR = "1B5E20"
OUT_FONT_COLOR = "B71C1C"
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN_BORDER = _make_border("thin")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def _fmt_time_12(h: int, m: int) -> str:
    ampm = "AM" if h < 12 else "PM"
    h12 = h % 12 or 12
    return f"{h12}:{m:02d} {ampm}"


def _first_entry_per_day(records: list) -> list:
    """Keep only the earliest record for each calendar day."""
    by_day: dict = {}
    for r in records:
        d = r["datetime"].date()
        if d not in by_day or r["datetime"] < by_day[d]["datetime"]:
            by_day[d] = r
    return sorted(by_day.values(), key=lambda x: x["datetime"])


def generate_excel(teachers: list, settings: dict, national_holidays: set, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active

    # Determine sheet name from records
    if teachers and teachers[0]["records"]:
        d = teachers[0]["records"][0]["datetime"]
        ws.title = d.strftime("%b-%y")
    else:
        ws.title = "Attendance"

    # Build header text
    sh, sm = map(int, settings["school_time"].split(":"))
    stah, stam = map(int, settings["staff_timing"].split(":"))
    relax = settings["relaxation_minutes"]
    total_m = stah * 60 + stam + relax
    cut_h, cut_m = total_m // 60, total_m % 60

    header_text = (
        f" School Time {_fmt_time_12(sh, sm)}, "
        f"Staff Timing {_fmt_time_12(stah, stam)}, "
        f"and after {relax} minutes relaxation {_fmt_time_12(cut_h, cut_m)}  "
    )

    # Column widths
    col_widths = {1: 9, 2: 22, 3: 22, 4: 10, 5: 3, 6: 9, 7: 22, 8: 22, 9: 10}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    row = 1
    for i in range(0, len(teachers), 2):
        t1 = teachers[i]
        t2 = teachers[i + 1] if i + 1 < len(teachers) else None

        # ── School timing header ──────────────────────────────────────
        ws.merge_cells(f"A{row}:I{row}")
        c = ws.cell(row=row, column=1, value=header_text)
        c.fill = HEADER_FILL
        c.font = Font(color="FFFFFF", bold=True, size=11)
        c.alignment = LEFT
        ws.row_dimensions[row].height = 22
        row += 1

        # ── Column headers ────────────────────────────────────────────
        col_labels = ["User ID", "Name", "Date/Time", "Status"]
        for j, label in enumerate(col_labels):
            c = ws.cell(row=row, column=j + 1, value=label)
            c.fill = SUBHEADER_FILL
            c.font = WHITE_FONT
            c.alignment = CENTER
            c.border = THIN_BORDER
        ws.cell(row=row, column=5).fill = PatternFill("solid", fgColor="F5F5F5")
        if t2:
            for j, label in enumerate(col_labels):
                c = ws.cell(row=row, column=j + 6, value=label)
                c.fill = SUBHEADER_FILL
                c.font = WHITE_FONT
                c.alignment = CENTER
                c.border = THIN_BORDER
        ws.row_dimensions[row].height = 18
        row += 1

        # ── Data rows (one entry per day; extras ignored) ───────────────
        recs1 = _first_entry_per_day(t1["records"])
        recs2 = _first_entry_per_day(t2["records"]) if t2 else []
        late_dates1 = t1.get("late_dates", set())
        late_dates2 = t2.get("late_dates", set()) if t2 else set()
        max_rows = max(len(recs1), len(recs2))

        for idx in range(max_rows):
            alt_fill = ALT_ROW_FILL if idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

            if idx < len(recs1):
                r = recs1[idx]
                fill = LATE_ROW_FILL if r["datetime"].date() in late_dates1 else alt_fill
                status_color = IN_FONT_COLOR if r["status"].lower() == "in" else OUT_FONT_COLOR
                _write_record_row(ws, row, 1, r, fill, status_color)

            if idx < len(recs2) and t2:
                r = recs2[idx]
                fill = LATE_ROW_FILL if r["datetime"].date() in late_dates2 else alt_fill
                status_color = IN_FONT_COLOR if r["status"].lower() == "in" else OUT_FONT_COLOR
                _write_record_row(ws, row, 6, r, fill, status_color)

            row += 1

        # ── Blank separator ───────────────────────────────────────────
        row += 1

        # ── Summary stats ─────────────────────────────────────────────
        stats = [
            ("Right Time Arrival", t1["right_time"], t2["right_time"] if t2 else None, "1B5E20"),
            ("Late Arrival", t1["late"], t2["late"] if t2 else None, "E65100"),
            ("Leave", t1["leave"], t2["leave"] if t2 else None, "B71C1C"),
        ]
        for label, val1, val2, color in stats:
            ws.row_dimensions[row].height = 17
            for col_offset, val in [(1, val1), (6, val2)]:
                if col_offset == 6 and t2 is None:
                    continue
                if col_offset == 6 and val is None:
                    continue
                lc = ws.cell(row=row, column=col_offset + 1, value=label)
                lc.fill = SUMMARY_FILL
                lc.font = Font(bold=True, color=color)
                lc.alignment = LEFT
                lc.border = THIN_BORDER

                vc = ws.cell(row=row, column=col_offset + 2, value=val)
                vc.fill = SUMMARY_FILL
                vc.font = Font(bold=True, color=color, size=12)
                vc.alignment = CENTER
                vc.border = THIN_BORDER
            row += 1

        # ── Gap before next block ─────────────────────────────────────
        row += 2

    wb.save(output_path)


def _write_record_row(ws, row: int, col_start: int, record: dict, fill, status_color: str):
    uid_c = ws.cell(row=row, column=col_start, value=record["user_id"])
    uid_c.fill = fill
    uid_c.alignment = CENTER
    uid_c.border = THIN_BORDER

    name_c = ws.cell(row=row, column=col_start + 1, value=record["name"])
    name_c.fill = fill
    name_c.alignment = LEFT
    name_c.border = THIN_BORDER

    dt_c = ws.cell(row=row, column=col_start + 2, value=record["datetime"])
    dt_c.fill = fill
    dt_c.number_format = "DD-MMM-YYYY HH:MM:SS"
    dt_c.alignment = CENTER
    dt_c.border = THIN_BORDER

    st_c = ws.cell(row=row, column=col_start + 3, value=record["status"])
    st_c.fill = fill
    st_c.font = Font(color=status_color, bold=True)
    st_c.alignment = CENTER
    st_c.border = THIN_BORDER
